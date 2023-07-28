import random
import openpyxl
from datetime import datetime, timedelta
# Define the categories for each column
age_groups = ["25>", "25-35", "36-45", "46-55", "56-65"]
gender_choices = ["Male", "Female", "Prefer not to say"]
marital_status_choices = ["Single", "Married", "Divorced", "Prefer not to say"]

# Add the new categories and choices here
educational_background_choices = [
    "Study or hold BSc in Dentistry",
    "Study or hold MSc in Dentistry",
    "Study or hold Ph.D. in Dentistry",
]

smoking_choices = ["No", "Yes"]

sports_time_choices = [
    "I do not do sports",
    "Less than 2 hours",
    "2 to 4 hours",
    "4 to 6 hours",
    "6 to 8 hours",
    "More than 8 hours",
]

general_health_choices = ["Poor", "Fair", "Good", "Very good", "Excellent"]

chronic_disease_choices = ["No", "Yes"]

regular_medication_choices = ["No", "Yes"]

influenza_vaccine_2022_2023_choices = ["No", "Yes"]

influenza_vaccine_previous_seasons_choices = ["Never", "At least 1 time"]

covid_diagnosed_choices = ["No", "Yes"]

covid_severity_choices = ["Not had", "Mild", "Moderate", "Severe"]

family_covid_diagnosed_choices = ["No", "Yes"]

covid_information_source_choices = [
    "Television",
    "Radio",
    "Social media",
    "Scientific journals",
    "Friends",
    "Recognized international health websites",
    "Health specialists",
    "Governmental sources",
]

covid_vaccinated_choices = ["No", "Yes"]

covid_vaccine_received_choices = [
    "BioNTech, Pfizer vaccine",
    "Sputnik V vaccine",
    "Oxford, AstraZeneca vaccine",
    "Sinopharm BBIBP vaccine",
    "The Sinovac-CoronaVac",
    "Johnson & Johnson vaccine",
    "Cuban Abdala vaccine",
]

willing_to_get_vaccinated_choices = [
    "No",
    "Probably No",
    "Probably Yes",
    "Yes",
]

vaccine_choice_choices = [
    "BioNTech, Pfizer vaccine",
    "Sputnik V vaccine",
    "Oxford, AstraZeneca vaccine",
    "Sinopharm BBIBP vaccine",
    "The Sinovac-CoronaVac",
    "Johnson & Johnson vaccine",
    "Cuban Abdala vaccine",
]

vaccine_choice_reasons_choices = [
    "Country of origin",
    "Based on a medical expert’s advice",
    "Articles you have interacted with on social media",
    "Family and friends’ recommendation",
    "Following the steps of a life role model",
    "I did not get to choose the vaccine (my choices were limited)",
    "Others: Please mention",
]

vaccination_reasons_choices = [
    "To protect family and friends",
    "To protect myself",
    "To protect patients",
    "To return to normal activities (travels, concerts, celebrations)",
    "To not miss days of work",
    "To comply with health ministry recommendations",
    "To not wear masks anymore",
]

not_vaccination_reasons_choices = [
    "Lack of information about COVID-19 vaccine",
    "COVID-19 vaccine is unsafe",
    "Fear of adverse events",
    "Pharmaceutical companies influence decisions about vaccination policies",
    "Previous diagnosis of COVID-19",
    "Suboptimal protective efficacy",
    "Disagree with vaccinations",
    "COVID-19 is not a threatening disease",
]

# Create a list containing all the categories lists
categories = [
    age_groups,
    gender_choices,
    marital_status_choices,
    educational_background_choices,
    smoking_choices,
    sports_time_choices,
    general_health_choices,
    chronic_disease_choices,
    regular_medication_choices,
    influenza_vaccine_2022_2023_choices,
    influenza_vaccine_previous_seasons_choices,
    covid_diagnosed_choices,
    covid_severity_choices,
    family_covid_diagnosed_choices,
    covid_information_source_choices,
    covid_vaccinated_choices,
    covid_vaccine_received_choices,
    willing_to_get_vaccinated_choices,
    vaccine_choice_choices,
    vaccine_choice_reasons_choices,
    vaccination_reasons_choices,
    not_vaccination_reasons_choices,
]

# Number of rows and columns in the Excel file
num_rows = 157
num_columns = 23

# Define the date range for the timestamp
start_date = datetime(2023, 7, 20, 0, 0)
end_date = datetime(2023, 7, 28, 23, 59)

# Generate random data for the Excel file
data = []
time_interval = (end_date - start_date) / num_rows
current_time = start_date

for _ in range(num_rows):
    row_data = []
    # Generate the timestamp in sequential order within the date range
    timestamp_str = current_time.strftime("%Y/%m/%d %I:%M:%S %p GMT+3")
    row_data.append(timestamp_str)
    current_time += time_interval
    # Rest of the data generation (same as in your original code)
    for category in categories:
        row_data.append(random.choice(category))
    data.append(row_data)

# Apply the conditions for the specific columns
for row_data in data:
    sports_time_choice = row_data[6]
    if sports_time_choice in ["4 to 6 hours", "6 to 8 hours", "More than 8 hours"]:
        row_data[7] = random.choice(["Good", "Very good", "Excellent"])
    
    if row_data[12] == "No":
        row_data[13] = ""
    
    if row_data[16] == "No":
        row_data[17] = ""
        row_data[19] = ""
    
    if row_data[17] in ["Probably No", "No"]:
        row_data[20] = ""
    
    if row_data[17] in ["Probably Yes", "Yes"]:
        row_data[21] = ""

# Create the Excel workbook and add data to the worksheet
workbook = openpyxl.Workbook()
sheet = workbook.active

# Set the headers for each column
headers = [
    "Timestamp",
    "Age",
    "Gender",
    "Marital Status",
    "Educational Background",
    "Smoke or Consume Tobacco",
    "Time Spent on Sports",
    "General Health",
    "Chronic Disease",
    "Regular Medication",
    "Received Influenza Vaccination 2022-2023",
    "Received Influenza Vaccination in Previous Influenza Seasons",
    "Diagnosed with COVID-19 Infection",
    "Severity of COVID-19 Symptoms",
    "Friends or Family Diagnosed with COVID-19 Infection",
    "Main Source of Information about COVID-19 Vaccines",
    "Received COVID-19 Vaccine",
    "COVID-19 Vaccine Received",
    "Willing to Get Vaccinated",
    "Vaccine Choice",
    "Reasons for Vaccine Choice",
    "Reasons for Vaccination or Willingness to Get Vaccinated",
    "Reasons for Not Getting Vaccinated or Willingness Not to Get Vaccinated",
]

for col_idx in range(num_columns):
    sheet.cell(row=1, column=col_idx + 1, value=headers[col_idx])

# Fill the data into the Excel sheet
for row_idx, row_data in enumerate(data):
    # Add the timestamp to the first column
    sheet.cell(row=row_idx + 2, column=1, value=row_data[0])
    # Fill the rest of the data for the row (same as in your original code)
    for col_idx, cell_value in enumerate(row_data[1:]):
        sheet.cell(row=row_idx + 2, column=col_idx + 2, value=cell_value)


# Save the Excel file
workbook.save("generated_data.xlsx")
